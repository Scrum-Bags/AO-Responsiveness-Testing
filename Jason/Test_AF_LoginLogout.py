import unittest
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from AF_AdminLogin import AF_Login
import HTML_Reporting
import time
import random
import string
import openpyxl
import win32com.client as win32


class Test_LoginLogout(unittest.TestCase):

    def __init__(self, *args, **kwargs):
        super(Test_LoginLogout, self).__init__(*args, **kwargs)
        self.timestr = time.strftime("%Y-%m-%d--%I_%M_%S%p")
        self.reporter = HTML_Reporting.TestSuiteReporter(self.timestr, "D:\\TestingResources\\AlineFinancial\\TestResults\\", "Jason")
        self.screenshotPath = "D:\\TestingResources\\AlineFinancial\\TestResults\\.screenshots\\"
        self.path = "D:\\TestingResources\\AlineFinancial\\DataSheets\\InputData.xlsm"
        self.xl = win32.Dispatch("Excel.Application")
        #self.xl.Interactive = False
        #self.xl.Visible = False
        xlbook = self.xl.Workbooks.Open(self.path)
        self.xl.Application.Run("InputData.xlsm!Module1.ResetLastCell()")
        xlbook.Save()
        self.xl.Application.Quit()

    def setUp(self):
        options = Options()
        options.headless = True
        self.driver = webdriver.Chrome(options=options)
        #self.driver.set_window_size(400,400)
        #self.driver.get("http://uftcapstone-dev-admin.s3-website-us-east-1.amazonaws.com/login")

    def test_001_login(self):
        #setup
        path = self.path
        wb = openpyxl.load_workbook(path)
        ws = wb["TC_001"]
        reporter = self.reporter
        driver = self.driver
        #repeat test with each data row
        for row in ws.iter_rows(min_row=2):

            print("##############################################")

            #set driver size
            res = row[2].value.split("x")
            if len(res)==2:
                driver.set_window_size(int(res[0]), int(res[1]))
##            elif res[0]=="fullscreen":
##                driver.fullscreen_window()
##                driver.manage().window().maximize();
            elif res[0] is None:
                pass

            #print(driver.get_window_size())
            
            #set up reporter
            testID="TC_001" +"_"+ ''.join(random.choices(string.ascii_lowercase, k=5))
            reporter.addTestCase(testID, "JS_TC_001", "Login to Aline Financial - Admin")
            reporter[testID].reportEvent("Set resolution for testing",False,res)

            #login process
            loginObj = AF_Login(driver)
            loginObj.Launch_Login_Page()
            loginObj.AF_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            #time.sleep(3)
            loginObj.AF_logout(reporter[testID], self.screenshotPath)
            #time.sleep(5)

    def test_002_login_neg(self):
        #setup
        path = self.path
        wb = openpyxl.load_workbook(path)
        ws = wb["TC_002"]
        reporter = self.reporter
        driver = self.driver
        #repeat test with each data row
        for row in ws.iter_rows(min_row=2):

            print("##############################################")
            #set driver size
            res = row[2].value.split("x")
            if len(res)==2:
                driver.set_window_size(int(res[0]), int(res[1]))
##            elif res[0]=="fullscreen":
##                driver.fullscreen_window()
##                driver.manage().window().maximize();
            elif res[0] is None:
                pass
            
            #set up reporter
            testID="TC_002" +"_"+ ''.join(random.choices(string.ascii_lowercase, k=5))
            reporter.addTestCase(testID, "JS_TC_002", "Attempt to login to Aline Financial - Admin with bad credentials")
            #login process
            loginObj = AF_Login(driver)
            loginObj.Launch_Login_Page()
            loginObj.AF_bad_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            #time.sleep(3)
            #loginObj.AO_logout(reporter[testID], self.screenshotPath)
            #time.sleep(5)



    def tearDown(self):
        self.driver.close()
        

if __name__ == "__main__":
    unittest.main(warnings='ignore')
    #unittest.main()
