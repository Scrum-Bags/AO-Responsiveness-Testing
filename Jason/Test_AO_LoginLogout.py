import unittest
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from AO_LoginLogout import AO_Login
from AO_Nav import AO_Nav
from Outlook import Outlook_App
import HTML_Reporting
import time
import random
import string
import openpyxl
import win32com.client as win32

#https://stackoverflow.com/questions/17353213/init-for-unittest-testcase

##
##timestr = time.strftime("%Y-%m-%d--%I_%M_%S%p")
##reporter = HTML_Reporting.TestSuiteReporter(timestr, "D:\\TestingResources\\AutomationPracticeStore\\TestResults\\", "Jason")

class Test_OrderHistory(unittest.TestCase):

    def __init__(self, *args, **kwargs):
        super(Test_OrderHistory, self).__init__(*args, **kwargs)
        self.timestr = time.strftime("%Y-%m-%d--%I_%M_%S%p")
        self.reporter = HTML_Reporting.TestSuiteReporter(self.timestr, "D:\\TestingResources\\AdvantageOnline\\TestResults\\", "Jason")
        self.screenshotPath = "D:\\TestingResources\\AdvantageOnline\\TestResults\\.screenshots\\"
        self.path = "D:\\TestingResources\\AdvantageOnline\\DataSheets\\InputData.xlsm"
        self.xl = win32.Dispatch("Excel.Application")
        #self.xl.Interactive = False
        #self.xl.Visible = False
        xlbook = self.xl.Workbooks.Open(self.path)
        self.xl.Application.Run("InputData.xlsm!Module1.ResetLastCell()")
        xlbook.Save()
        self.xl.Application.Quit()

    def setUp(self):
        options = Options()
        options.headless = True
        self.driver = webdriver.Chrome(options=options)
        self.driver.get("https://www.advantageonlineshopping.com/#/")

##    def test_001_login(self):
##        #setup
##        path = self.path
##        wb = openpyxl.load_workbook(path)
##        ws = wb["TC_001"]
##        reporter = self.reporter
##        driver = self.driver
##        #repeat test with each data row
##        for row in ws.iter_rows(min_row=2):
##            #set up reporter
##            testID="TC_001" +"_"+ ''.join(random.choices(string.ascii_lowercase, k=5))
##            reporter.addTestCase(testID, "JS_TC_001", "Login to Advantage Online")
##            #login process
##            loginObj = AO_Login(driver)
##            loginObj.Launch_Login_Page()
##            loginObj.AO_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
##            #time.sleep(3)
##            loginObj.AO_logout(reporter[testID], self.screenshotPath)
##            #time.sleep(5)
##
##    def test_002_login_neg(self):
##        #setup
##        path = self.path
##        wb = openpyxl.load_workbook(path)
##        ws = wb["TC_002"]
##        reporter = self.reporter
##        driver = self.driver
##        #repeat test with each data row
##        for row in ws.iter_rows(min_row=2):
##            #set up reporter
##            testID="TC_002" +"_"+ ''.join(random.choices(string.ascii_lowercase, k=5))
##            reporter.addTestCase(testID, "JS_TC_002", "Attempt to login to Advantage Online with bad credentials")
##            #login process
##            loginObj = AO_Login(driver)
##            loginObj.Launch_Login_Page()
##            loginObj.AO_bad_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
##            #time.sleep(3)
##            #loginObj.AO_logout(reporter[testID], self.screenshotPath)
##            #time.sleep(5)
##
##    def test_003_remember_login(self):
##        #setup
##        path = self.path
##        wb = openpyxl.load_workbook(path)
##        ws = wb["TC_003"]
##        reporter = self.reporter
##        driver = self.driver
##        #repeat test with each data row
##        for row in ws.iter_rows(min_row=2):
##            #set up reporter
##            testID="TC_003" +"_"+ ''.join(random.choices(string.ascii_lowercase, k=5))
##            reporter.addTestCase(testID, "JS_TC_003", "Login to Advantage Online with the 'remember me' button enabled")
##            #login process
##            loginObj = AO_Login(driver)
##            loginObj.Launch_Login_Page()
##            loginObj.AO_login_remember_test(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
##

    def test_004_logout(self):
        #setup
        path = self.path
        wb = openpyxl.load_workbook(path)
        ws = wb["TC_004"]
        reporter = self.reporter
        driver = self.driver
        #repeat test with each data row
        for row in ws.iter_rows(min_row=2):
            print("##############################################")

            res = row[2].value.split("x")
            if len(res)==2:
                driver.set_window_size(int(res[0]), int(res[1]))
            elif res[0]=="fullscreen":
                driver.fullscreen_window()
                driver.manage().window().maximize();
            elif res[0] is None:
                pass
            
            #set up reporter
            testID="TC_004" +"_"+ ''.join(random.choices(string.ascii_lowercase, k=5))
            reporter.addTestCase(testID, "JS_TC_004", "Logout on different pages of Advantage Online")
            reporter[testID].reportEvent("Set resolution for testing",False,res)
            #login process
            loginObj = AO_Login(driver)
            loginObj.Launch_Login_Page()
            result = loginObj.AO_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            if result:
                loginObj.AO_logout(reporter[testID], self.screenshotPath)
            #time.sleep(5)
            #Go to account page
            loginObj.Launch_Login_Page()
            result = loginObj.AO_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            if result:
                navObj = AO_Nav(driver)
                navObj.GoTo_Account(reporter[testID], self.screenshotPath)
                loginObj.AO_logout(reporter[testID], self.screenshotPath)
            #time.sleep(5)
            #Go to orders page
            loginObj.Launch_Login_Page()
            result = loginObj.AO_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            if result:
                navObj = AO_Nav(driver)
                navObj.GoTo_Orders(reporter[testID], self.screenshotPath)
                loginObj.AO_logout(reporter[testID], self.screenshotPath)
            #ime.sleep(5)
            #Go to shopping cart page
            loginObj.Launch_Login_Page()
            result = loginObj.AO_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            if result:
                navObj = AO_Nav(driver)
                navObj.GoTo_ShoppingCart(reporter[testID], self.screenshotPath)
                loginObj.AO_logout(reporter[testID], self.screenshotPath)
            #time.sleep(5)
             #Go to About page
            loginObj.Launch_Login_Page()
            result = loginObj.AO_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            if result:
                navObj = AO_Nav(driver)
                navObj.GoTo_About(reporter[testID], self.screenshotPath)
                loginObj.AO_logout(reporter[testID], self.screenshotPath)
            #time.sleep(5)
             #Go to AOS page
            loginObj.Launch_Login_Page()
            result = loginObj.AO_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            if result:
                navObj = AO_Nav(driver)
                navObj.GoTo_AOS(reporter[testID], self.screenshotPath)
                loginObj.AO_logout(reporter[testID], self.screenshotPath)
            #time.sleep(5)
             #Go to Speakers page
            loginObj.Launch_Login_Page()
            result = loginObj.AO_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            if result:
                navObj = AO_Nav(driver)
                navObj.GoTo_Speakers(reporter[testID], self.screenshotPath)
                loginObj.AO_logout(reporter[testID], self.screenshotPath)
            #time.sleep(5)
            
##
##    def test_005_forgot_password(self):
##        #setup
##        path = self.path
##        wb = openpyxl.load_workbook(path)
##        ws = wb["TC_005"]
##        reporter = self.reporter
##        driver = self.driver
##        #repeat test with each data row
##        for row in ws.iter_rows(min_row=2):
##            #set up reporter
##            testID="TC_005" +"_"+ ''.join(random.choices(string.ascii_lowercase, k=5))
##            reporter.addTestCase(testID, "JS_TC_005", "Go through the Advantage Online forgot password process")
##            #login process
##            loginObj = AO_Login(driver)
##            loginObj.Launch_Login_Page()
##            loginObj.AO_forgot_password(reporter[testID], self.screenshotPath, row[0].value)


    def tearDown(self):
        self.driver.close()
        

if __name__ == "__main__":
    unittest.main(warnings='ignore')
    #unittest.main()
