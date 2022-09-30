import random
import unittest
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from SwagLabs.SLCommonFuctions import login, mytime
from TestSuiteReporter import TestSuiteReporter
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select


class TestCases(unittest.TestCase):
  
    def test_001_browse_store_pos(self):
        reporter = TestSuiteReporter("SwagLabs", "C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/", "Kimberly Modeste")
        wb = openpyxl.load_workbook("C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLBrowseStore"]      
      
        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')
        
        reporter.addTestCase("Filter A-Z", "TC001", "User will be browsing the SwagLabs website "+wsBrowseStore.cell(row = 3, column = 1).value)

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser, reporter,"Filter A-Z")
        
        filter = Select(browser.find_element(by=By.CLASS_NAME, value="product_sort_container"))
        val = wsBrowseStore.cell(row = 2, column = 1).value
        if not val is None:
            filter.select_by_value(wsBrowseStore.cell(row = 2, column = 1).value)
        else:
            val = ""

        reporter["Filter A-Z"].reportEvent(eventDescription="User will be changing the filter", warning=False, 
        dataString=val, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)

        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            num.remove(num[j])

        reporter["Filter A-Z"].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Number of Items Selected: {i}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)

        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        reporter["Filter A-Z"].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)
        
    def test_002_browse_store_pos(self):
        reporter = TestSuiteReporter("SwagLabs", "C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/", "Kimberly Modeste")
        wb = openpyxl.load_workbook("C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLBrowseStore"]      
      
        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')
        
        reporter.addTestCase("Filter Z-A", "TC002", "User will be browsing the SwagLabs website "+wsBrowseStore.cell(row = 5, column = 1).value)

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser, reporter,"Filter Z-A")
        
        filter = Select(browser.find_element(by=By.CLASS_NAME, value="product_sort_container"))
        val = wsBrowseStore.cell(row = 4, column = 1).value
        if not val is None:
            filter.select_by_value(wsBrowseStore.cell(row = 4, column = 1).value)
        else:
            val = ""

        reporter["Filter Z-A"].reportEvent(eventDescription="User will be changing the filter", warning=False, 
        dataString=val, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)

        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            num.remove(num[j])

        reporter["Filter Z-A"].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Number of Items Selected: {i}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)

        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        reporter["Filter Z-A"].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)
      
    def test_003_browse_store_pos(self):
        reporter = TestSuiteReporter("SwagLabs", "C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/", "Kimberly Modeste")
        wb = openpyxl.load_workbook("C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLBrowseStore"]      
      
        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')
        
        reporter.addTestCase("Filter lohi", "TC003", "User will be browsing the SwagLabs website "+wsBrowseStore.cell(row = 7, column = 1).value)

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser, reporter,"Filter lohi")
        
        filter = Select(browser.find_element(by=By.CLASS_NAME, value="product_sort_container"))
        val = wsBrowseStore.cell(row = 6, column = 1).value
        if not val is None:
            filter.select_by_value(wsBrowseStore.cell(row = 6, column = 1).value)
        else:
            val = ""

        reporter["Filter lohi"].reportEvent(eventDescription="User will be changing the filter", warning=False, 
        dataString=val, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)

        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            num.remove(num[j])

        reporter["Filter lohi"].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Number of Items Selected: {i}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)

        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        reporter["Filter lohi"].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)
      
    def test_004_browse_store_pos(self):
        reporter = TestSuiteReporter("SwagLabs", "C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/", "Kimberly Modeste")
        wb = openpyxl.load_workbook("C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLBrowseStore"]      
      
        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')
        
        reporter.addTestCase("Filter hilo", "TC004", "User will be browsing the SwagLabs website "+wsBrowseStore.cell(row = 9, column = 1).value)

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser, reporter,"Filter hilo")
        
        filter = Select(browser.find_element(by=By.CLASS_NAME, value="product_sort_container"))
        val = wsBrowseStore.cell(row = 8, column = 1).value
        if not val is None:
            filter.select_by_value(wsBrowseStore.cell(row = 8, column = 1).value)
        else:
            val = ""

        reporter["Filter hilo"].reportEvent(eventDescription="User will be changing the filter", warning=False, 
        dataString=val, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)

        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            num.remove(num[j])

        reporter["Filter hilo"].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Number of Items Selected: {i}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)

        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        reporter["Filter hilo"].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath="C:/Users/OWNER/OneDrive/Documents/UFTOne/reports/Images/img"+mytime(), imageEmbed=False)
     
unittest.main()

