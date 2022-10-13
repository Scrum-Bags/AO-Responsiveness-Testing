import unittest
import openpyxl
import random
userStr = "C:\\Users\\Owner"
userStr = userStr.replace('\\', '/')
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from SLCommonFunctionsJenkins import loginHeadless, login, mytime, rand
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select


class TestCases(unittest.TestCase):
  
    #Commented
    def atest_001_place_order_pos(self):
        wb = openpyxl.load_workbook(f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      
      
        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')

        r = 2
        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser)
        
        filter = Select(browser.find_element(by=By.CLASS_NAME, value="product_sort_container"))
        val = wsBrowseStore.cell(row = r, column = 1).value
        if not val is None:
            filter.select_by_value(wsBrowseStore.cell(row = r, column = 1).value)
        else:
            val = ""

        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            num.remove(num[j])

        # Add purchase here
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        WebDriverWait(browser, 60).until(expected_conditions.presence_of_element_located((By.ID, "checkout")))
        browser.find_element(By.ID, "checkout").click()

        WebDriverWait(browser, 60).until(expected_conditions.presence_of_element_located((By.ID, "first-name")))
        print(wsBrowseStore.cell(row = r, column = 2).value)
        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)

        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)

        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
    
        browser.find_element(By.ID, "continue").click()
        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()


        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

    #Commented
    def atest_002_place_order_blank(self):
        wb = openpyxl.load_workbook(f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      
      
        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')
        
        TCN = "BlankFilterA-Z"
        r = 4
        
        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser)

        filter = Select(browser.find_element(by=By.CLASS_NAME, value="product_sort_container"))
        val = wsBrowseStore.cell(row = r, column = 1).value
        if not val is None:
            filter.select_by_value(wsBrowseStore.cell(row = r, column = 1).value)
        else:
            val = ""

        # Add purchase here
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        try:
            browser.find_element(By.CLASS_NAME, "cart_item").click()
        except:
            ActualBehavior="Pass"
            TestStatus=False

        browser.find_element(By.ID, "checkout").click()

        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
       
        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)
       
        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
        

        browser.find_element(By.ID, "continue").click()
        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()


        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        
    def test_003_place_order(self):
        wb = openpyxl.load_workbook(f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      
      
        options = Options()
        options.headless = True
        browser = webdriver.Chrome(options=options)
        browser.get('https://www.saucedemo.com/')
        print(f"Browser Opened: {browser.title}")

        r = 2
       
        print(f"Starting to Login...")
        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            loginHeadless(browser)
        print(f"Login completed")

        print(f"Trying to add items to cart")
        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            val = inventory[num[j]].find_element(by=By.CLASS_NAME, value="inventory_item_name").text
            print(f"Added to cart: {val}")
            num.remove(num[j])

        print(f"Added {i} items to cart")


        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        try:
            browser.find_element(By.CLASS_NAME, "cart_item").click()
        except:
            ActualBehavior="Pass"
            TestStatus=False

        browser.find_element(By.ID, "checkout").click()
        print(f"Clicked checked out")

        #First Name
        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
        print(f"Added first name")

        #Last Name
        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)
       
        print(f"Added last name")

        #Zipcode
        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
        print(f"Added Zipcode")


        browser.find_element(By.ID, "continue").click()
        print(f"Clicked Continue")

        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()

        print(f"Finished checkout")
        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        print(f"Logged out")

    #Commented
    def atest_004_place_order(self):
        wb = openpyxl.load_workbook(f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      

        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')

        TCN = "SwagLabsResponsive"
        r = 2
        
        val = rand()
        browser.set_window_size(*val)
        print(f"Browser set to: {val}")

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser)

        val = rand()
        browser.set_window_size(*val)
        

        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            val = inventory[num[j]].find_element(by=By.CLASS_NAME, value="inventory_item_name").text
            print(f"Added to cart: {val}")
            num.remove(num[j])

       
        val = rand()
        browser.set_window_size(*val)
       
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        try:
            browser.find_element(By.CLASS_NAME, "cart_item").click()
        except:
            ActualBehavior="Pass"
            TestStatus=False

        browser.find_element(By.ID, "checkout").click()

        val = rand()
        browser.set_window_size(*val)
       
        #First Name
        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
        

        #Last Name
        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)

        #Zipcode
        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
       
        val = rand()
        browser.set_window_size(*val)
       
        browser.find_element(By.ID, "continue").click()
         
        val = rand()
        browser.set_window_size(*val)
        
        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()

        val = rand()
        browser.set_window_size(*val)
        
        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        val = rand()
        browser.set_window_size(*val)
    
    #Commented
    def atest_005_place_order(self):
        wb = openpyxl.load_workbook(f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      

        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')

        
        r = 2
        
        browser.set_window_size(360, 640)

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser)


        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            val = inventory[num[j]].find_element(by=By.CLASS_NAME, value="inventory_item_name").text
            print(f"Added to cart: {val}")
            num.remove(num[j])

       
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        try:
            browser.find_element(By.CLASS_NAME, "cart_item").click()
        except:
            ActualBehavior="Pass"
            TestStatus=False

        browser.find_element(By.ID, "checkout").click()

       
        #First Name
        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
        

        #Last Name
        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)
        

        #Zipcode
        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
       
        browser.find_element(By.ID, "continue").click()
        
    
        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()

        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

unittest.main()