import string
from time import sleep
import unittest
import openpyxl
import random
import os
import zipfile
from S3_Tool import upload_file
import TestSuiteReporter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from SLCommonFunctionsJenkins import loginHeadless, mytime, rand
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions


class TestCases(unittest.TestCase):
    
    # @classmethod
    def __init__(self, *args, **kwargs):
        super(TestCases, self).__init__(*args, **kwargs)

    @classmethod
    def setUpClass(cls):
        cls.imageFolders = []
        cls.timestr = "SwagLabsJenkins"+mytime()
        cls.reporter = TestSuiteReporter.TestSuiteReporter(cls.timestr, "./", "Kimberly Modeste")
        cls.screenshotPath = "../.screenshots/"
        cls.path = "../TestCasesExcel.xlsx"


    def test_001_login_logout(self):
        print(f"#########################################")
       
        wb = openpyxl.load_workbook(self.path)  
      
        option = Options()
        option.headless = True
        browser = webdriver.Chrome(options=option)
        browser.get('https://www.saucedemo.com/')
        print(f"Browser Opened: {browser.title}")
        
        TCN = "SwagLabsHeadlessLoginout_"+"".join(random.choices(string.ascii_lowercase, k=10))
        self.imageFolders.append(TCN)
        r = 2
        self.reporter.addTestCase(TCN, "TC001", "User will be loging in and out of the SwagLabs Website")

        print(f"Starting to Login...")
        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            loginHeadless(browser, self.reporter,TCN)
        print(f"Login completed")

        # Logout
        WebDriverWait(browser, 5).until(expected_conditions.element_to_be_clickable(browser.find_element(by=By.ID, value="react-burger-menu-btn")))
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        self.reporter[TCN].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=''+'img'+mytime(), imageEmbed=False)
        print(f"Logged out")
        print(f"#########################################")
        browser.close()
     
    def test_002_place_order(self):
        print(f"#########################################")
        wb = openpyxl.load_workbook(f"../TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      
      
        option = Options()
        option.headless = True
        browser = webdriver.Chrome(options=option)
        browser.get('https://www.saucedemo.com/')
        print(f"Browser Opened: {browser.title}")
        
        TCN = "SwagLabsHeadlessWebSize_"+"".join(random.choices(string.ascii_lowercase, k=10))
        self.imageFolders.append(TCN)
        r = 2
        self.reporter.addTestCase(TCN, "TC002", "User will be purchasing from the SwagLabs website in headless mode randomly changing the size of the screen")

        val = {3000, 1082}
        browser.set_window_size(*val)
        print(f"Browser set to: {val}")
        self.reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Browser set to: {val}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot,  
        imagePath=''+'img'+mytime(), imageEmbed=False)

        print(f"Starting to Login...")
        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            loginHeadless(browser, self.reporter,TCN)
        print(f"Login completed")

        print(f"Trying to add items to cart")
        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        arr = []
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            val = inventory[num[j]].find_element(by=By.CLASS_NAME, value="inventory_item_name").text
            print(f"Added to cart: {val}")
            arr.append(val)
            num.remove(num[j])

        print(f"Added {i} items to cart")
        self.reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Number of Items Selected: {i}; \nItems Selected: {str(arr)[1:-2]}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=''+'img'+mytime(), imageEmbed=False)


        ActualBehavior="Pass"
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        try:
            browser.find_element(By.CLASS_NAME, "cart_item").click()
        except:
            ActualBehavior="Fail"

       
        browser.find_element(By.ID, "checkout").click()
        print(f"Clicked checked out")

        #First Name
        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
        self.reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 2).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 2).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=''+'img'+mytime(), imageEmbed=False)
        print(f"Added first name")


        #Last Name
        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)
        self.reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 3).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 3).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=''+'img'+mytime(), imageEmbed=False)
        print(f"Added last name")

    
        #Zipcode
        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
        self.reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 4).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 4).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=''+'img'+mytime(), imageEmbed=False)
        print(f"Added Zipcode")

        
        browser.find_element(By.ID, "continue").click()
        self.reporter[TCN].reportStep(stepDescription=wsBrowseStore.cell(row = r+1, column = 5).value, expectedBehavior=wsBrowseStore.cell(row = r, column = 5).value, 
        actualBehavior=(ActualBehavior==wsBrowseStore.cell(row = r, column = 5).value), testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=''+'img'+mytime(), imageEmbed=False)
        print(f"Clicked Continue")

       
        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()

        print(f"Finished checkout")
        WebDriverWait(browser, 5).until(expected_conditions.element_to_be_clickable(browser.find_element(by=By.ID, value="react-burger-menu-btn")))
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        self.reporter[TCN].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=''+'img'+mytime(), imageEmbed=False)
        print(f"Logged out")
        print(f"#########################################")
  


    @classmethod
    def tearDownClass(cls):
        del cls.reporter
        zipObj = zipfile.ZipFile(cls.timestr+".zip", 'w')
        zipObj.write(cls.timestr + ".html")
        for folder in cls.imageFolders:
            for image in os.listdir("./.screenshots/"+folder+"/"):
                zipObj.write("./.screenshots/"+folder+"/"+image)
        zipObj.close()
        upload_file(cls.timestr+".zip","scrumbags-reports")


unittest.main()
