import openpyxl
from sys import path
import os
userStr = os.environ['USERPROFILE']
userStr = userStr.replace('\\', '/')
path.append(f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test")
from Kimberly import TestSuiteReporter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from datetime import datetime
from testcases import testcasevalues, testcaseDict




def time():
    now = datetime.now()
    current_time = now.strftime("%Y%m%d_%H%M%S")
    return current_time

def login(
    browser: webdriver,
    reporter:  TestSuiteReporter, 
    TCRN: str, 
    r: int
):
    wb = openpyxl.load_workbook(f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
    wslogin = wb["APLogin"]

    email = wslogin.cell(row = r, column = 1).value
    password = wslogin.cell(row = r, column = 2).value

    browser.find_element(by=By.CLASS_NAME, value="login").click()
    browser.find_element(by=By.ID, value='email').send_keys(email)
    if email == "":
        Warning = True
    else:
        Warning = False

    reporter[TCRN].reportEvent(eventDescription=wslogin.cell(row = r+1, column = 1).value, warning=Warning, 
    dataString=f"Email: {email}", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)

    browser.find_element(by=By.ID, value='passwd').send_keys(password)
    if password == "":
        Warning = True
    else:
        Warning = False
    
    p = ""
    for i in password:
        p += "*"
    reporter[TCRN].reportEvent(eventDescription=wslogin.cell(row = r+1, column = 2).value, warning=Warning, 
    dataString=f"Password: {p}", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)
    browser.find_element(by=By.ID, value='SubmitLogin').click()

    WebDriverWait(browser, 60).until(
        expected_conditions.presence_of_element_located((By.XPATH, "//a[@title='Information']")) or 
        expected_conditions.presence_of_element_located((By.CLASS_NAME, "alert.alert-danger"))
    )
    
    if expected_conditions.presence_of_element_located((By.XPATH, "//a[@title='Information']")):
        ActualBehavior = "Pass"
        TestStatus = True
    else:
        ActualBehavior = "Fail"
        TestStatus = False

    reporter[TCRN].reportStep(stepDescription=wslogin.cell(row = r+1, column = 3).value, expectedBehavior=wslogin.cell(row = r, column = 3).value, 
    actualBehavior=ActualBehavior, testStatus=TestStatus, dataString="", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot,
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)
    
def create_account(browser, email):
    browser.find_element(By.ID, 'email_create').send_keys(email)
    browser.find_element(By.ID, 'SubmitCreate').click()

def to_input(item, str):
    item.click()
    item.clear()
    item.send_keys(str)

def personal_info_update(
browser: webdriver, 
reporter: TestSuiteReporter, 
TCRN: str,
r: int
):
    wb = openpyxl.load_workbook(f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
    ws = wb["APPersonalInfo"]
 
    browser.find_element(by=By.XPATH, value="//a[@title='Information']").click()
    WebDriverWait(browser, 60).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, "radio-inline")))

    radiobox = browser.find_elements(by=By.CLASS_NAME, value="radio-inline")
    i = 0
    if not ws.cell(row = r, column = 1).value is None:
        i = ws.cell(row = r, column = 1).value
    check = radiobox[i].find_element(by=By.XPATH, value="//label[starts-with(@for,'id_gender')]//div[@class='radio']//span")
    check.click()

    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 1).value, warning=False, 
    dataString=f"Gender: {i}", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)


    to_input(browser.find_element(by=By.XPATH, value="//input[@id='firstname']"), ws.cell(row = r, column = 2).value)
    to_input(browser.find_element(by=By.XPATH, value="//input[@id='lastname']"), ws.cell(row = r, column = 3).value)
    to_input(browser.find_element(by=By.XPATH, value="//input[@id='email']"), ws.cell(row = r, column = 4).value)
    if ws.cell(row = r, column = 2).value == "" or ws.cell(row = r, column = 3).value == "" or ws.cell(row = r, column = 4).value == "":
        Warning = True
    else:
        Warning = False
    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 2).value, warning=Warning, 
    dataString=f"First Name: {ws.cell(row = r, column = 2).value}, Last Name: {ws.cell(row = r, column = 3).value}, Email: {ws.cell(row = r, column = 4).value}", 
    screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", 
    imageEmbed=False)


    month = Select(browser.find_element(by=By.XPATH, value="//*[@id='months']"))
    if not(ws.cell(row = r, column = 5).value == ""):
        month.select_by_index(int(ws.cell(row = r, column = 5).value))
    else:
        Warning = True

    day = Select(browser.find_element(by=By.XPATH, value="//*[@id='days']"))
    if not(ws.cell(row = r, column = 6).value == ""):
        day.select_by_value(str(ws.cell(row = r, column = 6).value))
    else:
        Warning = True

    year = Select(browser.find_element(by=By.XPATH, value="//*[@id='years']"))
    if not(ws.cell(row = r, column = 7).value == ""):
        year.select_by_value(str(ws.cell(row = r, column = 7).value))
    else:
        Warning = True
    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 5).value, warning=Warning, 
    dataString=f"Birthday: {ws.cell(row = r, column = 5).value}/{ws.cell(row = r, column = 6).value}/{ws.cell(row = r, column = 7).value}", 
    screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)
    

    to_input(browser.find_element(by=By.XPATH, value="//input[@id='old_passwd']"), ws.cell(row = r, column = 8).value)
    to_input(browser.find_element(by=By.XPATH, value="//input[@id='passwd']"), ws.cell(row = r, column = 9).value)
    to_input(browser.find_element(by=By.XPATH, value="//input[@id='confirmation']"), ws.cell(row = r, column = 10).value)
    if ws.cell(row = r, column = 8).value == "" or ws.cell(row = r, column = 9).value == "" or ws.cell(row = r, column = 10).value == "":
        Warning = True
    else:
        Warning = False
    op = ""
    for i in ws.cell(row = r, column = 8).value:
        op += "*"

    p = ""
    for i in ws.cell(row = r, column = 9).value:
       p += "*"

    np = ""
    for i in ws.cell(row = r, column = 10).value:
       np += "*"

    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 8).value, warning=Warning, 
    dataString=f"Old Password: {op}, Password: {p}, New Password: {np}", 
    screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)


    if ws.cell(row = r, column = 11).value:
        browser.find_element(by=By.ID, value="newsletter").click()
        Warning = False
    else:
     Warning = True

    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 11).value, warning=Warning, 
    dataString=f"Click Newsletter: {ws.cell(row = r, column = 11).value}", 
    screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)


    if ws.cell(row = r, column = 12).value:
        browser.find_element(by=By.ID, value="optin").click()
        Warning = False
    else:
        Warning = True
    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 12).value, warning=False, 
    dataString=f"Click Opt In: {ws.cell(row = r, column = 11).value}", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot,
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)
    
    browser.find_element(by=By.XPATH, value="//button[@name='submitIdentity']").click()
    WebDriverWait(browser, 60).until( expected_conditions.presence_of_element_located((By.CLASS_NAME, "alert")))

    try:
        browser.find_element(by=By.CLASS_NAME, value="alert.alert-success").click()
    except:
        ActualBehavior = "Fail"
        TestStatus = False
    else:
        ActualBehavior = "Pass"
        TestStatus = True
    reporter[TCRN].reportStep(stepDescription=ws.cell(row = r+1, column = 13).value, expectedBehavior=ws.cell(row = r, column = 12).value, 
    actualBehavior=ActualBehavior, testStatus=TestStatus, dataString="", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot,
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)

def update_address(
    browser: webdriver, 
    reporter: TestSuiteReporter,
    TCRN: str, 
    r: int
 ):
    wb = openpyxl.load_workbook(f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
    ws = wb["APUpdateAddress"]


    browser.find_element(by=By.XPATH, value="//a[@title='Addresses']").click()
    WebDriverWait(browser, 60).until(expected_conditions.presence_of_element_located((By.XPATH, "//a[@title='Update']")))
    browser.find_element(by=By.XPATH, value="//a[@title='Update']").click()

    #First Name and Last Name
    fn = ws.cell(row = r, column = 1).value
    ln = ws.cell(row = r, column = 2).value
    Warning = False
    if fn == "":
        Warning = True
    else:
        to_input(browser.find_element(by=By.ID, value='firstname'), fn)

    if ln == "":
        Warning = True
    else:
       to_input( browser.find_element(by=By.ID, value='lastname'), ln)
    reporter[TCRN].reportEvent(eventDescription= ws.cell(row = r+1, column = 1).value, warning=Warning, 
    dataString=f"First Name: {fn}, Last Name: {ln}", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)

    #Company
    company = ws.cell(row = r, column = 3).value
    to_input(browser.find_element(by=By.ID, value='company'), company)
    reporter[TCRN].reportEvent(eventDescription= ws.cell(row = r+1, column = 3).value, warning=Warning, 
    dataString=f"Company: {company}", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)
   
    #Address and Address2
    add1 = ws.cell(row = r, column = 4).value
    add2 = ws.cell(row = r, column = 5).value
    if add1 == "":
        Warning = True
    else:
        Warning = False
    to_input(browser.find_element(by=By.ID, value='address1'), add1)
    to_input(browser.find_element(by=By.ID, value='address2'), add2)
    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 4).value, warning=Warning, 
    dataString=f"Address: {add1}, Address2: {add2}", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)
   

    #City, State and Zipcode
    c = ws.cell(row = r, column = 6).value
    s = ws.cell(row = r, column = 7).value
    zip = ws.cell(row = r, column = 8).value
    Warning = False
    if c == "":
        Warning = True
    else:
        to_input(browser.find_element(by=By.ID, value='city'), c)
    if s == "":
        Warning = True
    else:
        state = Select(browser.find_element(by=By.XPATH, value="//*[@id='id_state']"))
        state.select_by_index(int(s))
    if zip == "":
        Warning = True
    else:
        to_input(browser.find_element(by=By.ID, value='postcode'), zip)
    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 6).value, warning=Warning, 
    dataString=f"City: {c}, State: {s}, Zipcode: {zip}", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)
   

    #Country
    con = ws.cell(row = r, column = 9).value
    Warning = False
    if con == "":
        Warning = True
    else:
        country = Select(browser.find_element(by=By.XPATH, value="//*[@id='id_country']"))
        country.select_by_index(int(con))
    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 9).value, warning=Warning, dataString=f"Country: {con}", 
    screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)


    #Home and Mobile Phone
    home = ws.cell(row = r, column = 10).value
    mobile = ws.cell(row = r, column = 11).value
    if home == "" and mobile == "":
        Warning = True
    else:
        Warning = False
    to_input(browser.find_element(by=By.ID, value='phone'), home)
    to_input(browser.find_element(by=By.ID, value='phone_mobile'), mobile)
    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 10).value, warning=Warning, 
    dataString=f"Home Phone: {home}, Mobile Phone: {mobile}",  screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)
    
    
    #Other Text
    s =  ws.cell(row = r, column = 12).value
    to_input(browser.find_element(by=By.ID, value='other'), s)
    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 12).value, warning=Warning, 
    dataString=f"Additional Information: {s}",  screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)
    

    #Alias
    a = ws.cell(row = r, column = 13).value
    if a == "":
        Warning = True
    else:
        Warning = False
        to_input(browser.find_element(by=By.ID, value='alias'), a)
    reporter[TCRN].reportEvent(eventDescription=ws.cell(row = r+1, column = 13).value, warning=Warning, 
    dataString=f"Alias: {a}",  screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot, 
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)

    browser.find_element(by=By.ID, value='submitAddress').click()
    WebDriverWait(browser, 60).until(
        lambda browser:
        expected_conditions.presence_of_element_located((By.XPATH, "//a[@title='Update']")) or
        expected_conditions.presence_of_element_located((By.CLASS_NAME, "alert"))
    )
    try:
        browser.find_element(by=By.CLASS_NAME, value="alert.alert-danger").click()
    except:
        ActualBehavior = "Pass"
        TestStatus = True
    else:
        ActualBehavior = "Fail"
        TestStatus = False

    reporter[TCRN].reportStep(stepDescription=ws.cell(row = r+1, column = 14).value, expectedBehavior=ws.cell(row = r, column = 14).value, 
    actualBehavior=ActualBehavior, testStatus=TestStatus, dataString="", screenshotCallback=browser.find_element(by=By.ID, value='page').screenshot,
    imagePath=f"{userStr}/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{time()}", imageEmbed=False)

def signout(browser):
    browser.find_element(by=By.CLASS_NAME, value="logout").click()


def loginlogout(
    reporter: TestSuiteReporter, 
    TC: str,
    TCRN: str, 
    reports: testcaseDict, 
    values: testcasevalues
):
    browser = webdriver.Chrome()
    browser.get('http://automationpractice.com/index.php')
    reporter.addTestCase(TCRN, TC, values['title'])

    if browser.find_element(by=By.CLASS_NAME, value="login"):
        login(browser, reporter, TCRN, values['login'], reports['loginDict'])
    signout(browser)
    browser.quit()
        

           