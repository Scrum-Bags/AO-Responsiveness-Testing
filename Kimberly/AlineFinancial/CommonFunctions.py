
import random
import openpyxl
import time


from sys import path
path.append("C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test")
from Kimberly import TestSuiteReporter
from sql import read_query, create_aline_sql_connection
from datetime import datetime
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions


def rand ():
    random.seed()
    w = random.randint(700, 4069)
    h = random.randint(700, 2160)
    return {w, h}

def mytime():
    now = datetime.now()
    current_time = now.strftime("%Y%m%d_%H%M%S")
    return current_time 

def to_input(item, str):
    item.click()
    item.clear()
    item.send_keys(str)


def customWait(
    browser: webdriver,
    elem1,
    elem2,
    count: int,
):

    i = 0
    count = count * 2
    count = count / 5
    while(i < count):
        try:
            browser.find_element(elem1[0], f"{elem1[1]}").click()
        except:
            i +=1
        else:
            i = 50

        try:
            browser.find_element(elem2[0], f"{elem2[1]}").click()
        except:
            i +=1
        else:
            i = 50
        time.sleep(5)

def login(
    browser: webdriver,
    reporter:  TestSuiteReporter, 
    TCRN: str, 
    r: int
):
    wb = openpyxl.load_workbook("C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
    wslogin = wb["AFLogin"]

    if  wslogin.cell(row = r, column = 1).value is None:
        username = ""
    else:
        username = wslogin.cell(row = r, column = 1).value
    if wslogin.cell(row = r, column = 2).value is None:
        password = ""
    else:
        password = wslogin.cell(row = r, column = 2).value

    customWait(browser, (By.CLASS_NAME, "btn.btn-outline-light.w-100.rounded-pill"), (By.CLASS_NAME, "btn.btn-outline-light.rounded-pill"), 60)

    WebDriverWait(browser, 60).until( expected_conditions.presence_of_element_located((By.CLASS_NAME, "app-modal-container")))

    browser.find_element(by=By.ID, value='username').send_keys(username)

    if username == "":
        Warning = True
    else:
        Warning = False
    reporter[TCRN].reportEvent(eventDescription=wslogin.cell(row = r+1, column = 1).value, warning=Warning, 
    dataString=f"Username: {username}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{mytime()}", imageEmbed=False)

   
    browser.find_element(by=By.ID, value='password').send_keys(password)
    if password == "":
        Warning = True
    else:
        Warning = False

    p = ""
    for i in password:
        p += "*"
    reporter[TCRN].reportEvent(eventDescription=wslogin.cell(row = r+1, column = 2).value, warning=Warning, 
    dataString=f"Password: {p}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{mytime()}", imageEmbed=False)
    
    browser.find_element(by=By.XPATH, value='//button[@type="submit"]').click()

    customWaitHeadless(browser, (By.XPATH, "//fa-icon[@icon='exclamation-circle']"), (By.CLASS_NAME, "nav-item.ng-star-inserted"), 60)
  
    if expected_conditions.presence_of_element_located((By.XPATH, "//a[@class='nav-link.p-3.nav-active']")):
        ActualBehavior = "Pass"
        TestStatus = True
    else:
        ActualBehavior = "Fail"
        TestStatus = False
    

    reporter[TCRN].reportStep(stepDescription=wslogin.cell(row = r+1, column = 3).value, expectedBehavior=wslogin.cell(row = r, column = 3).value, 
    actualBehavior=ActualBehavior, testStatus=TestStatus, dataString="", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot,
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{mytime()}", imageEmbed=False)

def transactionChecker(
    browser: webdriver,
    reporter: TestSuiteReporter,
    TCN: str,
    num: int, #Types
    r: int,
    c: int
):
    connection = create_aline_sql_connection()

    if num == 0:
        s = "date"
    elif num == 3:
        s = "amount"
    elif num == 4:
        s = "posted_balance"
    else:
        s = "description"

    wb = openpyxl.load_workbook("C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
    wsBrowseStore = wb["AFViewTransactions"]      

    if wsBrowseStore.cell(row = r, column = c).value is None:
        val = ""
    else:
        val = wsBrowseStore.cell(row = r, column = c).value
        if  type(val) == datetime:
            val = val.strftime("%m/%d/%Y")

    query = "Select * from alinedb.transaction Where "+s+" Like '%{$"+str(val)+"}%' and account_id = 829 or account_id = 830"
    res = read_query(connection, query)
    reporter[TCN].reportEvent(eventDescription="Searching the database for item", warning=False, dataString=f"There were {res.__len__()} results found", 
    screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot,
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

    customWait(browser, (By.XPATH, "/html/body/app-root/app-dashboard/div/div/app-summary/div/app-transactions-view/div[2]"),(By.XPATH, "/html/body/app-root/app-dashboard/div/div/app-summary/div/app-transactions-view") ,60)

    searchbar = browser.find_elements(By.XPATH, "//input[@name='searchTerm']")
    if searchbar[0].is_displayed():
        searchbar = searchbar[0]
    else:
        searchbar=searchbar[-1]

    to_input(searchbar, str(val))
    

    searchbutton = browser.find_elements(By.XPATH, "//*//app-search-transactions//*//button[@type='submit']")
    if searchbutton[0].is_displayed():
        searchbutton[0].click()
    else:
        searchbutton[-1].click()
  
    reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = c).value, warning=False, dataString=val, 
    screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot,
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
    
    clicker = browser.find_elements(By.CLASS_NAME, "pagination")
    if clicker[-1].is_displayed():
        clicker = clicker[int(len(clicker)/2):]
    else: 
        clicker = clicker[:int(len(clicker)/2)]

    count = 0
    ActualBehavior = "Pass"
    TestStatus = True

    if wsBrowseStore.cell(row = r, column = c+1).value == "Pass" and len(clicker) == 2:
        ActualBehavior = "Fail"
        TestStatus = False

    for x in range(int(len(clicker)-2)):
        
        webtable = browser.find_elements(By.XPATH, "//table//tbody")
        if searchbar[0].is_displayed():
            webtable = webtable[0]
        else:
            webtable=webtable[-1]


        items = webtable.find_elements(By.TAG_NAME, "tr")

        for y in range(len(items)):

            section = items[y].find_elements(By.TAG_NAME, "td")
            if not section[num].text.__contains__(str(val)):
                ActualBehavior = "Fail"
                TestStatus = False
                count += 1
        if not clicker[len(clicker)-1].is_displayed():
            clicker[len(clicker)-1].click()

    reporter[TCN].reportStep(stepDescription=wsBrowseStore.cell(row = r+1, column = c+1).value, 
    expectedBehavior=wsBrowseStore.cell(row = r, column = c+1).value, actualBehavior=ActualBehavior, testStatus=TestStatus, 
    dataString=f"There were {count} items that weren't in the search", 
    screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

    browser.refresh()


def customWaitHeadless(
    browser: webdriver,
    elem1,
    elem2,
    count: int,
):

    i = 0
    count = count * 2
    count = count / 5
    print("Prepairing to enter Wait, stand by")
    print("Loading", end="")
    while(i < count):
        try:
            browser.find_element(elem1[0], f"{elem1[1]}").click()
        except:
            i +=1
        else:
            i = 50

        try:
            browser.find_element(elem2[0], f"{elem2[1]}").click()
        except:
            i +=1
        else:
            i = 50
        time.sleep(5)
        print(".", end= "")
    print("")
    print("Wait Completed!")

def loginHeadless(
    browser: webdriver,
    reporter:  TestSuiteReporter, 
    TCRN: str, 
    r: int
):
    wb = openpyxl.load_workbook("C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
    wslogin = wb["AFLogin"]

    if  wslogin.cell(row = r, column = 1).value is None:
        username = ""
    else:
        username = wslogin.cell(row = r, column = 1).value
    if wslogin.cell(row = r, column = 2).value is None:
        password = ""
    else:
        password = wslogin.cell(row = r, column = 2).value
    
    print("Username and Password retrived")
    
# d-inline-block py-1 ms-1 fs-5 text-decoration-none text-nowrap align-items-center
# d-inline-block py-1 ms-1 fs-5 text-decoration-none text-nowrap align-items-center

    WebDriverWait(browser, 60).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME, "fixed-top")))
    if browser.find_element(By.CLASS_NAME, "navbar-nav.d-inline-flex.d-lg-none").is_displayed():
        browser.find_element(By.CLASS_NAME, "navbar-nav.d-inline-flex.d-lg-none").click()
        WebDriverWait(browser, 60).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME, "btn.btn-outline-light.w-100.rounded-pill")))
        browser.find_element(By.CLASS_NAME, "btn.btn-outline-light.w-100.rounded-pill").click()
    else:
        browser.find_element(By.CLASS_NAME, "btn.btn-outline-light.rounded-pill").click()
  
    print("Clicked the button to start login")

    WebDriverWait(browser, 60).until( expected_conditions.presence_of_element_located((By.CLASS_NAME, "app-modal-container")))

    browser.find_element(by=By.ID, value='username').send_keys(username)

    if username == "":
        Warning = True
    else:
        Warning = False
    reporter[TCRN].reportEvent(eventDescription=wslogin.cell(row = r+1, column = 1).value, warning=Warning, 
    dataString=f"Username: {username}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{mytime()}", imageEmbed=False)

    print("Entered username into field")
   
    browser.find_element(by=By.ID, value='password').send_keys(password)
    if password == "":
        Warning = True
    else:
        Warning = False

    p = ""
    for i in password:
        p += "*"
    reporter[TCRN].reportEvent(eventDescription=wslogin.cell(row = r+1, column = 2).value, warning=Warning, 
    dataString=f"Password: {p}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{mytime()}", imageEmbed=False)
    print("Entered password into field")

    browser.find_element(by=By.XPATH, value='//button[@type="submit"]').click()

    print("Clicks submit, now we wait for website to let us in")
    customWaitHeadless(browser, (By.XPATH, "//fa-icon[@icon='exclamation-circle']"), (By.CLASS_NAME, "bg-light.min-vh-100.ng-tns-c71-0"), 60)
  
    if expected_conditions.presence_of_element_located((By.XPATH, "//a[@class='nav-link.p-3.nav-active']")):
        ActualBehavior = "Pass"
        TestStatus = True
    else:
        ActualBehavior = "Fail"
        TestStatus = False
    

    reporter[TCRN].reportStep(stepDescription=wslogin.cell(row = r+1, column = 3).value, expectedBehavior=wslogin.cell(row = r, column = 3).value, 
    actualBehavior=ActualBehavior, testStatus=TestStatus, dataString="", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot,
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCRN}/img{mytime()}", imageEmbed=False)

def transactionCheckerHeadless(
    browser: webdriver,
    reporter: TestSuiteReporter,
    TCN: str,
    num: int, #Types
    r: int,
    c: int
):
    print("Creating a Connection with the sql database")
    connection = create_aline_sql_connection()

    if num == 0:
        s = "date"
    elif num == 3:
        s = "amount"
    elif num == 4:
        s = "posted_balance"
    else:
        s = "description"

    wb = openpyxl.load_workbook("C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/TestCasesExcel.xlsx")
    wsBrowseStore = wb["AFViewTransactions"]      

    if wsBrowseStore.cell(row = r, column = c).value is None:
        val = ""
    else:
        val = wsBrowseStore.cell(row = r, column = c).value
        if  type(val) == datetime:
            val = val.strftime("%m/%d/%Y")

    print("Querying for items that look 'Like' our data")
    query = "Select * from alinedb.transaction Where "+s+" Like '%{$"+str(val)+"}%' and account_id = 829 or account_id = 830"
    res = read_query(connection, query)
    print("Query Completed")  

    reporter[TCN].reportEvent(eventDescription="Searching the database for item", warning=False, dataString=f"There were {res.__len__()} results found", 
    screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot,
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

    customWait(browser, (By.XPATH, "/html/body/app-root/app-dashboard/div/div/app-summary/div/app-transactions-view/div[2]"),(By.XPATH, "/html/body/app-root/app-dashboard/div/div/app-summary/div/app-transactions-view") ,60)

    searchbar = browser.find_elements(By.XPATH, "//input[@name='searchTerm']")
    print(f"Finding correct search bar. Searchbar len: {len(searchbar)}")
    if searchbar[0].is_displayed():
        searchbar = searchbar[0]
    else:
        searchbar =searchbar[1]


    print(f"Finding correct search bar. Searchbar len: {len(searchbar)}")
    to_input(s, str(val))
    print("Entering the search term")

    searchbutton = browser.find_elements(By.XPATH, "//*//app-search-transactions//*//button[@type='submit']")
    print(f"Finding correct search button. Searchbutton len: {len(searchbutton)}")
    if searchbutton[0].is_displayed():
        searchbutton[0].click()
    else:
        searchbutton[1].click()
    
    print("Clicking the enter button")
  
    reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = c).value, warning=False, dataString=val, 
    screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot,
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
    
    clicker = browser.find_elements(By.CLASS_NAME, "pagination")
    if clicker[-1].is_displayed():
        clicker = clicker[int(len(clicker)/2):]
    else: 
        clicker = clicker[:int(len(clicker)/2)]
    print("Finding the page clicker and finding the next page button")
    
    count = 0
    ActualBehavior = "Pass"
    TestStatus = True

    if wsBrowseStore.cell(row = r, column = c+1).value == "Pass" and len(clicker) == 2:
        ActualBehavior = "Fail"
        TestStatus = False

    print("Checking in all the results whether we got the correct item using the correct search method")
    for x in range(int(len(clicker)-2)):
        
        webtable = browser.find_elements(By.XPATH, "//table//tbody")
        if searchbar[0].is_displayed():
            webtable = webtable[0]
        else:
            webtable=webtable[-1]
        items = webtable.find_elements(By.TAG_NAME, "tr")

        for y in range(len(items)):
            section = items[y].find_elements(By.TAG_NAME, "td")
            if not section[num].text.__contains__(str(val)):
                ActualBehavior = "Fail"
                TestStatus = False
                count += 1
        if not clicker[len(clicker)-1].is_displayed():
            clicker[len(clicker)-1].click()

    reporter[TCN].reportStep(stepDescription=wsBrowseStore.cell(row = r+1, column = c+1).value, 
    expectedBehavior=wsBrowseStore.cell(row = r, column = c+1).value, actualBehavior=ActualBehavior, testStatus=TestStatus, 
    dataString=f"There were {count} items that weren't in the search", 
    screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
    imagePath=f"C:/Users/OWNER/OneDrive/Documents/UFTOne/tests/selenium/Test/Kimberly/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

    browser.refresh()
    print("Refreshing the page")



