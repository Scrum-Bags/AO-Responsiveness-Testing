import unittest
import openpyxl
import random
userStr = ".."
from .. import TestSuiteReporter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from SLCommonFunctionsJenkins import loginHeadless, login, mytime, rand
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select


class TestCases(unittest.TestCase):
  
    #Commented
    def atest_001_place_order_pos(self):
        reporter = TestSuiteReporter("SwagLabs", f"{userStr}/Reports", "Kimberly Modeste")
        wb = openpyxl.load_workbook(f"../TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      
      
        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')
        
        TCN = "PosFilterA-Z"
        r = 2
        reporter.addTestCase(TCN, "TC001", "User will be purchasing from the SwagLabs website "+wsBrowseStore.cell(row = r+1, column = 1).value)

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser, reporter,TCN)
        
        filter = Select(browser.find_element(by=By.CLASS_NAME, value="product_sort_container"))
        val = wsBrowseStore.cell(row = r, column = 1).value
        if not val is None:
            filter.select_by_value(wsBrowseStore.cell(row = r, column = 1).value)
        else:
            val = ""

        reporter[TCN].reportEvent(eventDescription="User will be changing the filter", warning=False, 
        dataString=val, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            num.remove(num[j])

        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Number of Items Selected: {i}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)


        # Add purchase here
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        WebDriverWait(browser, 60).until(expected_conditions.presence_of_element_located((By.ID, "checkout")))
        browser.find_element(By.ID, "checkout").click()

        WebDriverWait(browser, 60).until(expected_conditions.presence_of_element_located((By.ID, "first-name")))
        print(wsBrowseStore.cell(row = r, column = 2).value)
        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 2).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 2).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 3).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 3).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 4).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 4).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        browser.find_element(By.ID, "continue").click()
        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()


        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        reporter[TCN].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
    
    #Commented
    def atest_002_place_order_blank(self):
        reporter = TestSuiteReporter("SwagLabs", f"{userStr}/Reports", "Kimberly Modeste")
        wb = openpyxl.load_workbook(f"../TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      
      
        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')
        
        TCN = "BlankFilterA-Z"
        r = 4
        reporter.addTestCase(TCN, "TC002", "User will be purchasing from the SwagLabs website "+wsBrowseStore.cell(row = r+1, column = 1).value)

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser, reporter,TCN)
        
        filter = Select(browser.find_element(by=By.CLASS_NAME, value="product_sort_container"))
        val = wsBrowseStore.cell(row = r, column = 1).value
        if not val is None:
            filter.select_by_value(wsBrowseStore.cell(row = r, column = 1).value)
        else:
            val = ""

        reporter[TCN].reportEvent(eventDescription="User will be changing the filter", warning=False, 
        dataString=val, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
     
        # Add purchase here
        ActualBehavior="Pass"
        TestStatus=True
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        try:
            browser.find_element(By.CLASS_NAME, "cart_item").click()
        except:
            ActualBehavior="Pass"
            TestStatus=False

        browser.find_element(By.ID, "checkout").click()

        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 2).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 2).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 3).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 3).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 4).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 4).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        reporter[TCN].reportStep(stepDescription="User should finish the checkout", 
        expectedBehavior="Fail", actualBehavior=ActualBehavior, testStatus=TestStatus, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        browser.find_element(By.ID, "continue").click()
        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()


        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        reporter[TCN].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

    def test_003_place_order(self):
        reporter = TestSuiteReporter.TestSuiteReporter("SwagLabs", f"{userStr}/Reports", "Kimberly Modeste")
        wb = openpyxl.load_workbook(f"../TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      
      
        options = Options()
        options.headless = True
        browser = webdriver.Chrome(options=options)
        browser.get('https://www.saucedemo.com/')
        print(f"Browser Opened: {browser.title}")
        
        TCN = "SwagLabsHeadless"
        r = 2
        reporter.addTestCase(TCN, "TC003", "User will be purchasing from the SwagLabs website "+wsBrowseStore.cell(row = r+1, column = 1).value+ " in headless mode")

        print(f"Starting to Login...")
        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            loginHeadless(browser, reporter,TCN)
        print(f"Login completed")

        print(f"Trying to add items to cart")
        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            val = inventory[num[j]].find_element(by=By.CLASS_NAME, value="inventory_item_name").text
            print(f"Added to cart: {val}")
            num.remove(num[j])

        print(f"Added {i} items to cart")
        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Number of Items Selected: {i}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)


        ActualBehavior="Pass"
        TestStatus=True
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        try:
            browser.find_element(By.CLASS_NAME, "cart_item").click()
        except:
            ActualBehavior="Pass"
            TestStatus=False

        browser.find_element(By.ID, "checkout").click()
        print(f"Clicked checked out")

        #First Name
        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 2).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 2).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
        print(f"Added first name")

        #Last Name
        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 3).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 3).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
        print(f"Added last name")

        #Zipcode
        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 4).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 4).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
        print(f"Added Zipcode")


        browser.find_element(By.ID, "continue").click()
        reporter[TCN].reportStep(stepDescription="User should finish the checkout", 
        expectedBehavior="Fail", actualBehavior=ActualBehavior, testStatus=TestStatus, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
        print(f"Clicked Continue")

        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()

        print(f"Finished checkout")
        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        reporter[TCN].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
        print(f"Logged out")

    #Commented
    def atest_004_place_order(self):
        reporter = TestSuiteReporter.TestSuiteReporter("SwagLabs", f"{userStr}/Reports", "Kimberly Modeste")
        wb = openpyxl.load_workbook(f"../TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      

        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')

        TCN = "SwagLabsResponsive"
        r = 2
        reporter.addTestCase(TCN, "TC004", "User will be purchasing from the SwagLabs website "+wsBrowseStore.cell(row = r+1, column = 1).value+" with responsive testing")

        val = rand()
        browser.set_window_size(*val)
        print(f"Browser set to: {val}")

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser, reporter,TCN)

        val = rand()
        browser.set_window_size(*val)
        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Browser set to: {val}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)



        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            val = inventory[num[j]].find_element(by=By.CLASS_NAME, value="inventory_item_name").text
            print(f"Added to cart: {val}")
            num.remove(num[j])

        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Number of Items Selected: {i}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        val = rand()
        browser.set_window_size(*val)
        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Browser set to: {val}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        ActualBehavior="Pass"
        TestStatus=True
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        try:
            browser.find_element(By.CLASS_NAME, "cart_item").click()
        except:
            ActualBehavior="Pass"
            TestStatus=False

        browser.find_element(By.ID, "checkout").click()

        val = rand()
        browser.set_window_size(*val)
        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Browser set to: {val}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        #First Name
        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 2).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 2).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)


        #Last Name
        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 3).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 3).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)


        #Zipcode
        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 4).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 4).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
       
        val = rand()
        browser.set_window_size(*val)
        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Browser set to: {val}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        browser.find_element(By.ID, "continue").click()
        reporter[TCN].reportStep(stepDescription="User should finish the checkout", 
        expectedBehavior="Fail", actualBehavior=ActualBehavior, testStatus=TestStatus, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
       
        val = rand()
        browser.set_window_size(*val)
        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Browser set to: {val}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()

        val = rand()
        browser.set_window_size(*val)
        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Browser set to: {val}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        reporter[TCN].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        val = rand()
        browser.set_window_size(*val)
        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Browser set to: {val}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
      
    #Commented
    def atest_005_place_order(self):
        reporter = TestSuiteReporter.TestSuiteReporter("SwagLabs", f"{userStr}/Reports", "Kimberly Modeste")
        wb = openpyxl.load_workbook(f"../TestCasesExcel.xlsx")
        wsBrowseStore = wb["SLPlaceOrder"]      

        browser = webdriver.Chrome()
        browser.get('https://www.saucedemo.com/')

        TCN = "SwagLabsResponsive"
        r = 2
        reporter.addTestCase(TCN, "TC004", "User will be purchasing from the SwagLabs website "+wsBrowseStore.cell(row = r+1, column = 1).value+" with a set responsive testing")

        
        browser.set_window_size(360, 640)

        if browser.find_element(by=By.CLASS_NAME, value="login-box"):
            login(browser, reporter,TCN)


        inventory = browser.find_element(by=By.CLASS_NAME, value="inventory_list").find_elements(by=By.CLASS_NAME, value="inventory_item")

        num = list(range(0, inventory.__len__()))
        random.seed()
        i = random.randint(1, 5)
        for x in range(i):
            j = random.randint(0, num.__len__()-1)
            inventory[num[j]].find_element(by=By.TAG_NAME, value="button").click()
            val = inventory[num[j]].find_element(by=By.CLASS_NAME, value="inventory_item_name").text
            print(f"Added to cart: {val}")
            num.remove(num[j])

        reporter[TCN].reportEvent(eventDescription="User selected items", warning=False, 
        dataString=f"Number of Items Selected: {i}", screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)

        ActualBehavior="Pass"
        TestStatus=True
        browser.find_element(By.CLASS_NAME, "shopping_cart_link").click()
        try:
            browser.find_element(By.CLASS_NAME, "cart_item").click()
        except:
            ActualBehavior="Pass"
            TestStatus=False

        browser.find_element(By.ID, "checkout").click()

       
        #First Name
        browser.find_element(By.ID, "first-name").send_keys(wsBrowseStore.cell(row = r, column = 2).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 2).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 2).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)


        #Last Name
        browser.find_element(By.ID, "last-name").send_keys(wsBrowseStore.cell(row = r, column = 3).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 3).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 3).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)


        #Zipcode
        browser.find_element(By.ID, "postal-code").send_keys(wsBrowseStore.cell(row = r, column = 4).value)
        reporter[TCN].reportEvent(eventDescription=wsBrowseStore.cell(row = r+1, column = 4).value, warning=False, 
        dataString=wsBrowseStore.cell(row = r, column = 4).value, screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
       
     
        browser.find_element(By.ID, "continue").click()
        reporter[TCN].reportStep(stepDescription="User should finish the checkout", 
        expectedBehavior="Fail", actualBehavior=ActualBehavior, testStatus=TestStatus, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)
       
    
        browser.find_element(By.ID, "finish").click()
        browser.find_element(By.ID, "back-to-products").click()

        # Logout
        browser.find_element(by=By.ID, value="react-burger-menu-btn").click()
        buttonList = browser.find_element(by=By.CLASS_NAME, value="bm-item-list").find_elements(by=By.TAG_NAME, value="a")
        WebDriverWait(browser, 60).until(expected_conditions.element_to_be_clickable(buttonList[2]))
        buttonList[2].click()

        reporter[TCN].reportStep(stepDescription="User should click the hamburger and logout", 
        expectedBehavior="Pass", actualBehavior="Pass", testStatus=True, dataString="", 
        screenshotCallback=browser.find_element(by=By.TAG_NAME, value='body').screenshot, 
        imagePath=f"{userStr}/.screenshots/{TCN}/img{mytime()}", imageEmbed=False)


unittest.main()
