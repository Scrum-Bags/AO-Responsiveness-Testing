from selenium import webdriver
from TestSuiteReporter import TestSuiteReporter
import logging
from typing import Callable, Union
import mysql.connector
from mysql.connector import Error
import random, string, time, inspect
import pathlib
from openpyxl import load_workbook

import logging
import boto3
from botocore.exceptions import ClientError
import os


def upload_file(file_name, bucket, object_name=None):
    """Upload a file to an S3 bucket

    :param file_name: File to upload
    :param bucket: Bucket to upload to
    :param object_name: S3 object name. If not specified then file_name is used
    :return: True if file was uploaded, else False
    """

    # If S3 object_name was not specified, use file_name
    if object_name is None:
        object_name = os.path.basename(file_name)

    # Upload the file
    s3_client = boto3.client('s3')
    try:
        response = s3_client.upload_file(file_name, bucket, object_name)
    except ClientError as e:
        logging.error(e)
        return False
    return True

if __name__ == '__main__':
    #usage_demo()
    #main()
    #hello_s3()
    print(upload_file("2022-10-12--04_50_10PM.html", "scrumbags-reports"))
    pass

def report_event_and_log(driver, message: str):
    driver.reporter[driver.testID].reportEvent(message, False, "")
    func = inspect.currentframe().f_back.f_code
    path = pathlib.PurePath(func.co_filename)
    logging.getLogger(driver.loggingID).info("{%s in %s:%i} - %s" % (
        func.co_name,
        path.name,
        func.co_firstlineno,
        message
    ))

def log_wrapper(driver, message):
    func = inspect.currentframe().f_back.f_code
    path = pathlib.PurePath(func.co_filename)
    logging.getLogger(driver.loggingID).info("{%s in %s:%i} - %s" % (
        func.co_name,
        path.name,
        func.co_firstlineno,
        message
    ))

def load_excel_sheet(driver, rowName, file, sheetName):
    wb = load_workbook(filename = file, data_only=True)
    sheet = wb[sheetName]
    driver.data = {}
    for row in sheet.rows:
        if row[0].value == rowName:
            for i in range (2, 22):
                driver.data[sheet.cell(column=i, row=1).value] = sheet.cell(column=i, row=row[0].row).value
            report_event_and_log(driver, "Loaded excel data")

def excel_get_rows(file, sheetName):
    wb = load_workbook(filename = file, data_only=True)
    sheet = wb[sheetName]
    return sheet.max_row

def check_for_responsive(driver):
    log_wrapper(driver, "Checking for responsive webpage")
    if driver.get_window_size()['width'] < 550:
        driver.responsive = True
        log_wrapper(driver, "Detected responsive")
    else:
        driver.responsive = False
        log_wrapper(driver, "Didn't detect responsive")


#SQL Utilites
def create_aline_sql_connection(driver):
    connection = None
    try:
        connection = mysql.connector.connect(
            host='uftcapstone-db.c1ddjzxizuua.us-east-1.rds.amazonaws.com',
            user='uftcapstone',
            passwd='!A&8vYOKSUO&X9Zt',
            database='alinedb'
        )
        log_wrapper(driver, "MySQL Database connection sucessful")
    except Error as err:
        log_wrapper(driver, f"Error: '{err}'")

    return connection

def read_query(driver, connection, query):
    cursor = connection.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        return result
    except Error as err:
        log_wrapper(driver, f"Error: '{err}'")

def execute_query(driver, connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        connection.commit()
        log_wrapper(driver, "Query successful")
    except Error as err:
        log_wrapper(driver, f"Error '{err}'")

def find_and_update_email(driver, email):
    connection = create_aline_sql_connection(driver)
    query = "SELECT * FROM applicant WHERE email='" + email + "'"
    results = read_query(driver, connection, query)
    if len(results) > 0:
        random_email = ''.join(random.choices(string.ascii_lowercase, k=12)) + '@' + ''.join(random.choices(string.ascii_lowercase, k=12)) + '.com'
        update_str = "UPDATE applicant SET email='" + random_email + "' WHERE email='" + email + "'"
        execute_query(driver, connection, update_str)
